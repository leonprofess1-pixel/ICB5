import openpyxl
from openpyxl.styles import NamedStyle
from datetime import datetime, timedelta
import sys
import io

# Windows 콘솔의 인코딩 문제를 해결하기 위해 출력 인코딩을 UTF-8로 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def format_date_in_sheet(sheet):
    """'입력값' 시트의 K열과 L열의 날짜 형식을 변환합니다."""
    date_style = NamedStyle(name='custom_date', number_format='YYYY-MM-DD')
    
    if date_style.name not in sheet.parent.style_names:
        sheet.parent.add_named_style(date_style)

    for row in sheet.iter_rows(min_row=2):
        for cell in [row[10], row[11]]:
            if cell.value and not isinstance(cell.value, datetime):
                try:
                    date_val = None
                    if isinstance(cell.value, str) and len(cell.value) == 8:
                        date_val = datetime.strptime(cell.value, '%Y%m%d')
                    elif isinstance(cell.value, (int, float)):
                        date_str = str(int(cell.value))
                        if len(date_str) == 8:
                            date_val = datetime.strptime(date_str, '%Y%m%d')
                    
                    if date_val:
                        cell.value = date_val
                        cell.style = date_style
                except (ValueError, TypeError):
                    continue

def find_employee_row_map(sheet, emp_id_col_idx):
    """사번을 키로, 행 번호를 값으로 하는 맵을 생성합니다."""
    emp_map = {}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        emp_id = row[emp_id_col_idx - 1]
        if emp_id:
            emp_map[str(emp_id).strip()] = row_idx
    return emp_map

def find_next_empty_cell_col(sheet, row_idx, start_col, end_col):
    """지정된 행과 열 범위에서 첫 번째 빈 셀의 열 인덱스를 찾습니다."""
    for col_idx in range(start_col, end_col + 1):
        if sheet.cell(row=row_idx, column=col_idx).value is None:
            return col_idx
    return None

def main():
    """메인 실행 함수"""
    file_path = r'연차사용내역\vac_Test.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"오류: '{file_path}' 파일을 찾을 수 없습니다.")
        return

    # 시트 로드
    input_sheet = None
    if "입력값" in workbook.sheetnames: input_sheet = workbook["입력값"]
    elif "Sheet1" in workbook.sheetnames: input_sheet = workbook["Sheet1"]
    if not input_sheet:
        print("오류: '입력값' 또는 'Sheet1' 시트를 찾을 수 없습니다."); return

    vacation_sheet = None
    for name in workbook.sheetnames:
        if "연차사용현황" in name: vacation_sheet = workbook[name]; break
    if not vacation_sheet:
        print("오류: '연차사용현황'이 포함된 시트를 찾을 수 없습니다."); return

    exclusion_sheet = None
    if "연도 연차계산" in workbook.sheetnames:
        exclusion_sheet = workbook["연도 연차계산"]
    if not exclusion_sheet:
        print("오류: '연도 연차계산' 시트를 찾을 수 없습니다."); return
        
    print("정보: 휴일 제외 목록을 로드합니다.")
    exclusion_dates = set()
    for row_cells in exclusion_sheet['N5':'N40']:
        for cell in row_cells:
            if cell.value and isinstance(cell.value, datetime):
                exclusion_dates.add(cell.value.date())
    print(f"정보: 총 {len(exclusion_dates)}개의 제외할 날짜를 로드했습니다.")

    print(f"정보: '{input_sheet.title}' 시트의 데이터를 처리합니다.")
    format_date_in_sheet(input_sheet)
    emp_row_map = find_employee_row_map(vacation_sheet, emp_id_col_idx=5)
    
    processed_count = 0
    for row in input_sheet.iter_rows(min_row=2, max_row=input_sheet.max_row):
        progress_status = row[4].value
        emp_id = row[6].value
        leave_type_val = row[9].value
        start_date = row[10].value
        end_date = row[11].value

        if progress_status != "처리완료" or not all([emp_id, leave_type_val, start_date, end_date]):
            continue

        target_row_idx = emp_row_map.get(str(emp_id).strip())
        if not target_row_idx:
            continue
        
        dates_to_write = []
        if start_date.date() == end_date.date():
            dates_to_write.append(start_date)
        else:
            delta = end_date - start_date
            for i in range(delta.days + 1):
                dates_to_write.append(start_date + timedelta(days=i))

        leave_type_str = str(leave_type_val)
        target_range = None
        is_half_day = False
        if "연차" in leave_type_str:
            target_range = (22, 46)
        elif "반차" in leave_type_str:
            target_range = (47, 96)
            is_half_day = True
        else:
            continue

        for date_to_process in dates_to_write:
            weekday = date_to_process.weekday() # 월요일=0, 토요일=5, 일요일=6
            date_str_formatted = date_to_process.strftime('%Y-%m-%d')
            
            # [수정] 휴일 목록 또는 주말(토, 일)인 경우 건너뛰기
            if date_to_process.date() in exclusion_dates:
                print(f"정보: {date_str_formatted}은(는) 휴일 목록에 있어 건너뜁니다.")
                continue
            if weekday in [5, 6]:
                print(f"정보: {date_str_formatted}은(는) 주말(토/일)이므로 건너뜁니다.")
                continue

            value_to_insert = date_to_process
            if is_half_day:
                date_str = date_to_process.strftime("%m/%d")
                if "오전" in leave_type_str: value_to_insert = f"{date_str}(전)"
                elif "오후" in leave_type_str: value_to_insert = f"{date_str}(후)"
                else: value_to_insert = f"{date_str}(반차)"
            
            next_col = find_next_empty_cell_col(vacation_sheet, target_row_idx, target_range[0], target_range[1])
            if next_col:
                previous_cell_value = None
                if (target_range[0] == 22 and next_col > 22) or (target_range[0] == 47 and next_col > 47):
                    previous_cell_value = vacation_sheet.cell(row=target_row_idx, column=next_col - 1).value
                
                is_duplicate = False
                if isinstance(previous_cell_value, datetime) and isinstance(value_to_insert, datetime):
                    if previous_cell_value.date() == value_to_insert.date(): is_duplicate = True
                elif previous_cell_value == value_to_insert: is_duplicate = True

                if not is_duplicate:
                    vacation_sheet.cell(row=target_row_idx, column=next_col, value=value_to_insert)
                    processed_count += 1
                else:
                    print(f"정보: 행 {target_row_idx}에서 중복된 값 '{value_to_insert}'의 입력을 건너뜁니다.")

    print(f"정보: 총 {processed_count}개의 신규 데이터를 처리하여 입력했습니다.")

    print("\n정보: '입력값' 시트를 정리합니다.")
    input_sheet['V2'] = "작업 완료"
    if input_sheet.max_row > 1:
        input_sheet.delete_rows(2, input_sheet.max_row - 1)
        print("정보: '입력값' 시트의 헤더 아래 모든 데이터 행을 삭제했습니다.")

    try:
        workbook.save(file_path)
        print(f"\n작업 완료: 모든 변경사항이 원본 파일 '{file_path}'에 저장되었습니다.")
    except PermissionError:
        print(f"\n오류: '{file_path}'을(를) 저장할 수 없습니다. 파일이 다른 프로그램에서 열려있는지 확인하세요.")
    except Exception as e:
        print(f"파일 저장 중 예상치 못한 오류 발생: {e}")


if __name__ == "__main__":
    main()
